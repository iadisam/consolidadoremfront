import openpyxl
import os
import re

def consolidar_excels():
    archivo_base = 'SA_26_V1.2.xlsm'
    archivo_salida = 'SA_26_CONSOLIDADO.xlsm'
    
    # Obtener archivos .xlsm (excluyendo base y salida)
    archivos_a_sumar = [f for f in os.listdir('.') if f.endswith('.xlsm') 
                        and f != archivo_base and f != archivo_salida]
    
    print(f"Cargando plantilla base: {archivo_base}")
    wb_final = openpyxl.load_workbook(archivo_base, keep_vba=True)
    
    # Identificar todas las pestañas que empiezan con 'A' en la plantilla base
    # Esto incluirá A01, A11a, A19b, A30AR, etc.
    pestanas_objetivo = [sheet for sheet in wb_final.sheetnames if sheet.startswith('A')]
    print(f"Pestañas detectadas para consolidar: {', '.join(pestanas_objetivo)}")

    for nombre_archivo in archivos_a_sumar:
        print(f"Procesando: {nombre_archivo}")
        # data_only=True para obtener el resultado de fórmulas de los archivos de origen
        wb_datos = openpyxl.load_workbook(nombre_archivo, data_only=True)
        
        for nombre_hoja in pestanas_objetivo:
            if nombre_hoja in wb_datos.sheetnames:
                ws_datos = wb_datos[nombre_hoja]
                ws_final = wb_final[nombre_hoja]
                
                # Recorrer celdas que contienen datos en el archivo de origen
                for row in ws_datos.iter_rows():
                    for cell in row:
                        # Solo procesar si hay un valor numérico
                        if isinstance(cell.value, (int, float)):
                            celda_destino = ws_final[cell.coordinate]
                            
                            # CRÍTICO: Solo sumar si la celda en la BASE no es una fórmula
                            # Las celdas con fórmulas se calcularán solas en Excel al abrir el archivo
                            if celda_destino.data_type != 'f':
                                valor_actual = celda_destino.value if isinstance(celda_destino.value, (int, float)) else 0
                                celda_destino.value = valor_actual + cell.value

        wb_datos.close()

    print(f"Guardando archivo consolidado: {archivo_salida}")
    wb_final.save(archivo_salida)
    print("¡Finalizado con éxito!")

if __name__ == "__main__":
    consolidar_excels()